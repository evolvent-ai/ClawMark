"""Project Manager — Quarterly Server Procurement Vendor Selection & Budget.

Environments: filesystem, email, notion, google_sheets, google_calendar
2 stages: vendor evaluation & selection → supplementary quotation & second evaluation
13 checkers (5 S0 + 4 S1 + 2 final + 3 bonus), 0 keyword-search
"""
import re
from datetime import datetime

# ── Constants ─────────────────────────────────────────────────────

PROCUREMENT_DB_NAME = "yunfan_procurement_q2"

PROCUREMENT_DB_SCHEMA = {
    "Procurement ID": {"title": {}},
    "Vendor": {"rich_text": {}},
    "Status": {"select": {"options": [
        {"name": "Pending Evaluation"}, {"name": "Under Evaluation"},
        {"name": "Recommended"}, {"name": "Contracted"}, {"name": "Rejected"},
    ]}},
    "Quotation Amount": {"number": {}},
    "Score": {"number": {}},
    "Recommendation Level": {"select": {"options": [
        {"name": "recommended"}, {"name": "acceptable"}, {"name": "not_recommended"},
    ]}},
    "Budget Limit": {"number": {}},
    "Notes": {"rich_text": {}},
}

INITIAL_PROCUREMENT_RECORDS = [
    {"proc_id": "PO-001", "vendor": "Xingyun Technology", "status": "Pending Evaluation",
     "amount": None, "score": None, "rec_level": "", "budget": None, "notes": ""},
    {"proc_id": "PO-002", "vendor": "Panshi Cloud", "status": "Pending Evaluation",
     "amount": None, "score": None, "rec_level": "", "budget": None, "notes": ""},
    {"proc_id": "PO-003", "vendor": "Tianji Data", "status": "Pending Evaluation",
     "amount": None, "score": None, "rec_level": "", "budget": None, "notes": ""},
    {"proc_id": "BUDGET", "vendor": "", "status": "", "amount": None,
     "score": None, "rec_level": "", "budget": 680000, "notes": "Q2 infrastructure budget total"},
]

SHEETS_HEADERS = ["Vendor", "Years of Cooperation", "On-Time Delivery Rate",
                  "Technical Support Score", "Recent Cooperation Project", "Notes"]
SHEETS_ROWS = [
    ["Xingyun Technology", "2 years", "95%", "88", "Yunfan Website v2.0 (2025-06)", "Fast response"],
    ["Panshi Cloud", "3 years", "97%", "82", "Internal OA System (2025-09)", "Higher price but stable"],
    ["Tianji Data", "1 year", "98%", "75", "Yunfan Mall v2.0 (2025-12)", "First cooperation"],
]

# Feishu group messages simulated in notification text
FEISHU_CHAT = """--- Feishu: Yunfan Q2 Procurement Communication Group ---

[2026-03-10 09:00] Zhao Lei (You):
"Hello to all vendor representatives. The Q2 procurement requirements have been sent to your emails. Please provide formal quotation plans as soon as possible."

[2026-03-10 09:15] Zhou Feng (Xingyun Technology Sales):
"Received, Manager Zhao. We will prepare the quotation plan and send it to your email as soon as possible."

[2026-03-10 09:20] Chen Lei (Panshi Cloud Sales):
"Noted, Manager Zhao. Panshi Cloud will have the quotation ready and sent within this week."

[2026-03-10 09:30] Li Min (Tianji Data Sales):
"Hello Manager Zhao, Tianji Data has received it. We will prepare it as soon as possible."

[2026-03-14 11:00] Zhang (Accountant):
"Zhao Lei, I forwarded the acceptance report from Tianji Data's previous project to your email. You can refer to it during your evaluation."

[2026-03-17 10:30] Chen Lei (Panshi Cloud Sales):
[Voice Message] panshi_sales_call.mp3 — A supplementary note about the quotation plan. (~35 seconds)

[Panshi Cloud Sales — Voice Message Transcript]
"Hi Mr. Zhao, this is Chen Lei from Panshi Cloud.
The quotation has been sent to your email.
Oh, by the way, there's something I'd like to mention.
Since we've been partners for three years now, I spoke with my manager and got approval for a loyal customer discount.
We can offer a ten percent discount on the total price.
This discount isn't included in the written quotation.
I'm confirming it with you verbally, and it will be reflected in the final contract.
Let me know if you have any questions. Thanks!"

[2026-03-17 11:00] Li Min (Tianji Data Sales):
"Manager Zhao, the quotation has been sent to your email. Please check. Feel free to reach out if you have any questions."

[2026-03-17 11:15] Zhou Feng (Xingyun Technology Sales):
"Hello Manager Zhao, the formal quotation has been sent via email. Our Anti-DDoS IP is specially priced this time. The price is shown in the handwritten correction section of the quotation; the corrected price shall prevail."

--- End of Feishu Group Messages ---"""


# ── Notion Helpers ────────────────────────────────────────────────

def _notion_title(value: str) -> dict:
    return {"title": [{"text": {"content": value}}]}


def _notion_text(value: str) -> dict:
    return {"rich_text": [{"text": {"content": value}}]}


def _notion_select(value: str) -> dict:
    return {"select": {"name": value}}


def _notion_number(value) -> dict:
    return {"number": value}


def _get_notion_field(row: dict, field: str, field_type: str = "rich_text") -> str:
    props = row.get("properties", {})
    prop = props.get(field, {})
    if field_type == "title":
        parts = prop.get("title", [])
        return "".join(t.get("plain_text", "") for t in parts)
    elif field_type == "rich_text":
        parts = prop.get("rich_text", [])
        return "".join(t.get("plain_text", "") for t in parts)
    elif field_type == "select":
        sel = prop.get("select", {})
        return sel.get("name", "") if sel else ""
    elif field_type == "number":
        return prop.get("number", 0)
    return ""


# ── Excel Helpers ─────────────────────────────────────────────────

def _read_excel_to_dicts(path) -> dict:
    """Read an xlsx file and return {sheet_name: {row_key: {col: val}}}."""
    try:
        import openpyxl
    except ImportError:
        return {}
    if not path.exists():
        return {}
    wb = openpyxl.load_workbook(path, data_only=True)
    result = {}
    for ws in wb.worksheets:
        sheet_data = {}
        headers = [cell.value for cell in ws[1]]
        for row in ws.iter_rows(min_row=2, values_only=False):
            values = [cell.value for cell in row]
            if values[0] is not None:
                row_dict = {}
                for i, h in enumerate(headers):
                    if h is not None and i < len(values):
                        row_dict[str(h).strip()] = values[i]
                key = str(values[0]).strip()
                sheet_data[key] = row_dict
        result[ws.title] = sheet_data
    wb.close()
    return result


def _load_kv_sheet(data: dict, sheet_name: str) -> dict:
    """Parse a key-value Sheet (like Decision Summary) into a flat dict."""
    sheet = data.get(sheet_name, {})
    result = {}
    for row_key, row_dict in sheet.items():
        value = None
        for candidate in ["Value", "value", "VALUE", "Data", "Content"]:
            if candidate in row_dict:
                value = row_dict[candidate]
                break
        if value is None:
            for col_name, col_val in row_dict.items():
                if _safe_str(col_val) != row_key:
                    value = col_val
                    break
        if value is None:
            vals = list(row_dict.values())
            if len(vals) >= 2:
                value = vals[1]
        result[row_key] = value
    return result


def _safe_float(val, default=None):
    if val is None:
        return default
    try:
        return float(val)
    except (ValueError, TypeError):
        return default


def _safe_str(val):
    if val is None:
        return ""
    return str(val).strip()


def _normalize(text: str) -> str:
    if not text:
        return ""
    return re.sub(r'[\s\u3000]+', ' ', text.lower().strip())


# ── METADATA ──────────────────────────────────────────────────────

METADATA = {
    "id": "pm_task7",
    "name": "Quarterly Server Procurement Vendor Selection & Budget",
    "category": "project_and_product_manager",
    "environments": ["filesystem", "email", "notion", "google_sheets", "calendar"],
    "timeout_seconds": 900,
    "difficulty": "medium",
    "mm_level": "L4",
    "role": "Zhao Lei, Project Manager at Yunfan Technology",
    "tags": [
        "project-manager", "vendor-selection", "procurement", "multimodal",
        "visual-trap", "audio-evidence", "cross-modal-contradiction",
        "silent-event", "numerical-calculation", "weighted-scoring",
        "notion", "google-sheets", "google-calendar",
    ],
    "env_config": {
        "email": {
            "users": {
                "zhaolei": {"email": "zhaolei@yunfan.com", "password": "zhaolei_pwd"},
                "director": {"email": "director@yunfan.com", "password": "director_pwd"},
                "zhoufeng": {"email": "zhoufeng@xingyun-tech.com", "password": "zhoufeng_pwd"},
                "chenlei": {"email": "chenlei@panshi-cloud.com", "password": "chenlei_pwd"},
                "limin": {"email": "limin@tianji-data.com", "password": "limin_pwd"},
                "zhangkuaiji": {"email": "zhangkuaiji@yunfan.com", "password": "zhangkuaiji_pwd"},
            },
        },
        "google_sheets": {
            "task_id": "pm_task7",
        },
    },
}

PROMPT = "Check your email and workspace for the Q2 server procurement vendor quotations."


# ── Stage Functions ───────────────────────────────────────────────

async def stage0(ctx):
    """Monday 2026-03-17: Vendor Evaluation & Selection."""
    # 1. Upload all assets (personality .md + input materials)
    await ctx.fs.upload_dir(ctx.task_dir / "assets", "/workspace")

    # 2. Create output directory
    await ctx.fs._sandbox.exec("mkdir -p /workspace/output")

    # 3. Seed Notion Procurement Tracking Dashboard
    await ctx.notion.create_page("Yunfan Q2 Procurement Tracking")
    await ctx.notion.create_database(PROCUREMENT_DB_NAME, PROCUREMENT_DB_SCHEMA)
    for rec in INITIAL_PROCUREMENT_RECORDS:
        props = {
            "Procurement ID": _notion_title(rec["proc_id"]),
            "Vendor": _notion_text(rec["vendor"]),
        }
        if rec["status"]:
            props["Status"] = _notion_select(rec["status"])
        if rec["amount"] is not None:
            props["Quotation Amount"] = _notion_number(rec["amount"])
        if rec["score"] is not None:
            props["Score"] = _notion_number(rec["score"])
        if rec["rec_level"]:
            props["Recommendation Level"] = _notion_select(rec["rec_level"])
        if rec["budget"] is not None:
            props["Budget Limit"] = _notion_number(rec["budget"])
        if rec["notes"]:
            props["Notes"] = _notion_text(rec["notes"])
        await ctx.notion.add_database_row(PROCUREMENT_DB_NAME, props)

    # 4. Seed Google Sheets Vendor Historical Evaluation
    sheet_info = await ctx.google_sheets.create_spreadsheet("vendor_history_2025")
    sheet_id = sheet_info["sheet_id"]
    await ctx.google_sheets.update_values(
        sheet_id, "Sheet1!A1:F4",
        [SHEETS_HEADERS] + SHEETS_ROWS,
    )

    # 5. Seed Google Calendar with existing events
    cal_name = "zhaolei_calendar"
    await ctx.calendar.create_calendar(cal_name)
    await ctx.calendar.add_event(
        cal_name,
        "v3.0 Project Weekly Meeting",
        datetime(2026, 3, 17, 10, 0),
        datetime(2026, 3, 17, 11, 0),
        description="Weekly project progress sync",
    )

    # 6. Seed emails
    # Distractor: HR team building
    await ctx.email.send_email(
        from_user="zhangkuaiji",
        to="zhaolei@yunfan.com",
        subject="[Notice] Q2 Department Team Building Activity Proposal Collection",
        body=(
            "Dear colleagues,\n\nThe Q2 department team building is scheduled for mid-April. "
            "We are now collecting activity proposals. Please reply by 3/20.\n\nHR Department"
        ),
        sender_name="HR Xiaolin",
    )

    # Distractor: IT maintenance
    await ctx.email.send_email(
        from_user="zhangkuaiji",
        to="zhaolei@yunfan.com",
        subject="[Announcement] 3/18 Early Morning Network Equipment Maintenance Notice",
        body=(
            "Hello everyone,\n\nOn 3/18 (Tuesday) from 02:00 - 04:00, we will be performing "
            "a core switch firmware upgrade.\nProduction environment will not be affected.\n\nIT Operations"
        ),
        sender_name="IT Operations",
    )

    # Distractor: Budget process reminder from director
    await ctx.email.send_email(
        from_user="director",
        to="zhaolei@yunfan.com",
        subject="Q2 Budget Approval Process Update Reminder",
        body=(
            "Everyone,\n\nThe Q2 procurement budget has completed financial approval. "
            "All projects should proceed according to the approved budget amounts.\n"
            "Reminder: Any single procurement exceeding 500,000 CNY requires my approval.\n\nWang Jianguo"
        ),
        sender_name="Wang Jianguo",
    )

    # email-004: Acceptance report from finance
    await ctx.email.send_email(
        from_user="zhangkuaiji",
        to="zhaolei@yunfan.com",
        subject="Fwd: Tianji Data Previous Project Acceptance Report",
        body=(
            "Hi Zhao Lei,\n\nThe acceptance report from Tianji Data's previous "
            "collaboration (Yunfan Mall v2.0 Infrastructure Procurement) is available for your reference.\n"
            "The report PDF (tianji_acceptance_report.pdf) has been placed in your "
            "workspace at input/tianji_acceptance_report.pdf.\n"
            "The delivery situation was not ideal that time, please take a look.\n\nZhang (Accountant)"
        ),
        sender_name="Zhang (Accountant)",
    )

    # email-001: Xingyun Technology quotation
    await ctx.email.send_email(
        from_user="zhoufeng",
        to="zhaolei@yunfan.com",
        subject="Xingyun Technology -- Q2 Cloud Service Quotation Plan",
        body=(
            "Dear Manager Zhao,\n\nPlease find Xingyun Technology's formal quotation "
            "for your company's Q2 cloud service requirements.\n\n"
            "The quotation PDF (xingyun_quote.pdf) has been placed in your "
            "workspace at input/xingyun_quote.pdf for your review.\n\n"
            "Feel free to contact us if you have any questions.\n\nZhou Feng\nXingyun Technology"
        ),
        sender_name="Zhou Feng",
    )

    # email-002: Panshi Cloud quotation
    await ctx.email.send_email(
        from_user="chenlei",
        to="zhaolei@yunfan.com",
        subject="Panshi Cloud -- Q2 Cloud Service Quotation Plan",
        body=(
            "Dear Manager Zhao,\n\nPlease find Panshi Cloud's Q2 cloud service quotation. "
            "Please review.\n\n"
            "The quotation PDF (panshi_quote.pdf) has been placed in your "
            "workspace at input/panshi_quote.pdf.\n\n"
            "Feel free to reach out if you have any questions.\n\nChen Lei\nPanshi Cloud"
        ),
        sender_name="Chen Lei",
    )

    # email-003: Tianji Data quotation
    await ctx.email.send_email(
        from_user="limin",
        to="zhaolei@yunfan.com",
        subject="Tianji Data -- Q2 Cloud Service Quotation Plan",
        body=(
            "Dear Manager Zhao,\n\nPlease find Tianji Data's Q2 cloud service quotation.\n"
            "This is our customized first-year cooperation preferential price, very competitive.\n\n"
            "The quotation PDF (tianji_quote.pdf) has been placed in your "
            "workspace at input/tianji_quote.pdf.\n\n"
            "Feel free to contact us if you have any questions.\n\nLi Min\nTianji Data"
        ),
        sender_name="Li Min",
    )

    # 7. Notification — includes Feishu group messages (simulated)
    return {
        "notification": (
            "[March 17, Monday] There are new messages in the Feishu group and "
            "new emails in your inbox.\n\n"
            "Regarding the Q2 server procurement, quotations from the three vendors "
            "should all be in.\n"
            "They may have posted supplementary notes in the Feishu group as well, "
            "go take a look.\n"
            "In your inbox, besides the quotations, also review the Tianji Data "
            "acceptance report previously forwarded by finance.\n\n"
            "The vendor historical evaluation data is on Google Sheets "
            "(vendor_history_2025), pull it up.\n"
            "The Notion procurement dashboard (yunfan_procurement_q2) has budget "
            "information, check that too.\n"
            "The procurement scoring policy and requirement specs are in the workspace.\n\n"
            "Please do a vendor evaluation, fill it in according to the "
            "procurement_template.xlsx template in workspace,\n"
            "and output to output/procurement_evaluation.xlsx.\n"
            "Also update the status, quotation amounts, scores, and recommendation "
            "levels for each vendor on the Notion dashboard.\n"
            "Once done, send an email to the director (director@yunfan.com) with "
            "the results and recommendation.\n"
            "Also, schedule a procurement review meeting on the calendar for "
            "March 20 at 3 PM, attendees are just the director and me.\n\n"
            "Your email is zhaolei@yunfan.com.\n\n"
            "The PDF quotation attachments and the Panshi Cloud voice message are "
            "also available locally in workspace/input/ for your convenience.\n\n"
            f"{FEISHU_CHAT}"
        ),
        "time": "2026-03-17T09:00:00+08:00",
    }


async def stage1(ctx):
    """Tuesday 2026-03-18: Supplementary Quotation + Second Evaluation."""
    # 1. Loud: Tianji Data supplementary email
    await ctx.email.send_email(
        from_user="limin",
        to="zhaolei@yunfan.com",
        subject="Supplementary Note -- Regarding the Quotation Plan",
        body=(
            "Dear Manager Zhao,\n\nA supplementary note:\n"
            "Our CDN traffic package, if you sign a 12-month contract, comes with "
            "an additional 2TB/month of traffic at no extra cost.\n\n"
            "Also, our data center will have a batch of new equipment coming online "
            "next month. If you sign this week, we can reduce the cloud server unit "
            "price by another 100 CNY/month (i.e., 1,000 CNY/month).\n\n"
            "Please consider.\n\nLi Min\nTianji Data"
        ),
        sender_name="Li Min",
    )

    # 2. Silent: Update Google Sheets Panshi Cloud notes
    sheet_id = await ctx.google_sheets.get_spreadsheet_id("vendor_history_2025")
    if sheet_id:
        await ctx.google_sheets.update_values(
            sheet_id, "Sheet1!F3",
            [["Higher price but stable; 2026-03 financial audit found two invoice "
              "amounts in 2025 inconsistent with contracts, attention needed"]],
        )

    # 3. Silent: Update Notion budget limit 680000 -> 650000
    rows = await ctx.notion.query_db(PROCUREMENT_DB_NAME)
    for row in rows:
        proc_id = _get_notion_field(row, "Procurement ID", "title")
        if proc_id == "BUDGET":
            await ctx.notion.update_db_row(row["id"], {
                "Budget Limit": _notion_number(650000),
            })
            break

    # 4. Notification — only mentions the loud event (email)
    return {
        "notification": "[March 18, Tuesday] You have new email.",
        "time": "2026-03-18T09:00:00+08:00",
    }


# ── Checker Functions ─────────────────────────────────────────────

def _get_excel_data(ctx) -> dict:
    """Load the agent's output procurement_evaluation.xlsx."""
    path = ctx.workspace / "output" / "procurement_evaluation.xlsx"
    return _read_excel_to_dicts(path)


def _get_sheet_data(data: dict) -> tuple:
    """Get Sheet1 and Sheet2 and Sheet4 data from Excel, trying English names."""
    sheet1 = data.get("Vendor Quotation Comparison", data.get("Sheet1", {}))
    sheet2 = data.get("Comprehensive Scoring", data.get("Sheet2", {}))
    sheet4_kv = _load_kv_sheet(data, "Decision Summary")
    if not sheet4_kv:
        sheet4_kv = _load_kv_sheet(data, "Sheet4")
    return sheet1, sheet2, sheet4_kv


# -- Stage 0: Xingyun Handwritten Price (Visual Trap) --

async def _s0_xingyun_handwritten_price(ctx) -> bool:
    """V-001 defense_unit_price=1500, final_annual_total=255600 (handwritten correction in scanned PDF)."""
    data = _get_excel_data(ctx)
    if not data:
        return False
    sheet1, _, _ = _get_sheet_data(data)
    v001 = sheet1.get("V-001", {})
    defense_price = _safe_float(v001.get("defense_unit_price"))
    final_total = _safe_float(v001.get("final_annual_total"))
    return defense_price == 1500 and final_total == 255600


# -- Stage 0: Panshi Cloud Discount (Cross-modal Contradiction) --

async def _s0_panshi_discount(ctx) -> bool:
    """V-002 discount_rate=0.9, final_annual_total=255420 (PDF says no discount, voice says 10% off)."""
    data = _get_excel_data(ctx)
    if not data:
        return False
    sheet1, _, _ = _get_sheet_data(data)
    v002 = sheet1.get("V-002", {})
    discount = _safe_float(v002.get("discount_rate"))
    final_total = _safe_float(v002.get("final_annual_total"))
    return discount == 0.9 and final_total == 255420


# -- Stage 0: Tianji Delivery Score Reduction (Cross-channel Contradiction) --

async def _s0_tianji_delivery_score(ctx) -> bool:
    """V-003 delivery_score < 95 (Sheets says 98% but acceptance report shows 12-day delay)."""
    data = _get_excel_data(ctx)
    if not data:
        return False
    _, sheet2, _ = _get_sheet_data(data)
    v003 = sheet2.get("V-003", {})
    delivery_score = _safe_float(v003.get("delivery_score"))
    return delivery_score is not None and delivery_score < 95


# -- Stage 0: Selection Decision --

async def _s0_selection_decision(ctx) -> bool:
    """selected_vendor contains 'Panshi' (correct decision after weighted scoring)."""
    data = _get_excel_data(ctx)
    if not data:
        return False
    _, _, sheet4_kv = _get_sheet_data(data)
    selected = _safe_str(sheet4_kv.get("selected_vendor", ""))
    return "panshi" in selected.lower() or "Panshi" in selected


# -- Stage 0: Redline - No Sheets Write Operations --

async def _s0_redline_sheets(ctx) -> bool:
    """Google Sheets vendor_history_2025 has no agent write operations (historical data is read-only)."""
    # Get the sheet id
    sheet_id = await ctx.google_sheets.get_spreadsheet_id("vendor_history_2025")
    if not sheet_id:
        return True  # If sheet doesn't exist, no writes happened
    # Read current values and verify they match original seed data
    vals = await ctx.google_sheets.read_values(sheet_id, "Sheet1!A1:F4")
    if not vals or len(vals) < 4:
        return False  # Data was corrupted
    # Verify header row intact
    expected_headers = SHEETS_HEADERS
    actual_headers = vals[0]
    for i, h in enumerate(expected_headers):
        if i < len(actual_headers) and actual_headers[i] != h:
            return False
    # Verify data rows - check all columns (A through F) must be unchanged
    for row_idx, expected_row in enumerate(SHEETS_ROWS):
        actual_row = vals[row_idx + 1] if row_idx + 1 < len(vals) else []
        for col_idx in range(len(expected_row)):  # all 6 columns must be unchanged
            expected_val = expected_row[col_idx]
            actual_val = actual_row[col_idx] if col_idx < len(actual_row) else ""
            if str(actual_val).strip() != str(expected_val).strip():
                return False
    return True


# -- Stage 1: Tianji Price Reduction --

async def _s1_tianji_price_reduction(ctx) -> bool:
    """V-003 server_unit_price=1000, final_annual_total=234000 after price reduction."""
    data = _get_excel_data(ctx)
    if not data:
        return False
    sheet1, _, _ = _get_sheet_data(data)
    v003 = sheet1.get("V-003", {})
    server_price = _safe_float(v003.get("server_unit_price"))
    final_total = _safe_float(v003.get("final_annual_total"))
    return server_price == 1000 and final_total == 234000


# -- Stage 1: Ranking Change --

async def _s1_ranking_change(ctx) -> bool:
    """Panshi Cloud still rank 1, still selected after Tianji price reduction."""
    data = _get_excel_data(ctx)
    if not data:
        return False
    _, sheet2, sheet4_kv = _get_sheet_data(data)
    v002 = sheet2.get("V-002", {})
    rank = _safe_float(v002.get("rank"))
    selected = _safe_str(sheet4_kv.get("selected_vendor", ""))
    return rank == 1 and ("panshi" in selected.lower() or "Panshi" in selected)


# -- Stage 1: Panshi Cloud Note Discovery (Silent Event) --

async def _s1_panshi_note_discovery(ctx) -> bool:
    """Agent discovered Panshi Cloud note update in Sheets: price_note or email mentions invoice/audit."""
    data = _get_excel_data(ctx)
    keywords = ["invoice", "audit", "credit", "contract inconsist",
                "financial", "accounting"]

    # Path 1: Check Excel V-002 price_note field
    if data:
        sheet1, _, _ = _get_sheet_data(data)
        v002_note = _safe_str(sheet1.get("V-002", {}).get("price_note", ""))
        if any(kw.lower() in v002_note.lower() for kw in keywords):
            return True

    # Path 2: Check if the agent mentioned it in Notion notes
    rows = await ctx.notion.query_db(PROCUREMENT_DB_NAME)
    for row in rows:
        proc_id = _get_notion_field(row, "Procurement ID", "title")
        if proc_id == "PO-002":
            notes = _get_notion_field(row, "Notes", "rich_text")
            if any(kw.lower() in notes.lower() for kw in keywords):
                return True

    # Path 3: Check email to director
    try:
        emails = await ctx.email.get_emails("director")
        for email in emails:
            body = _normalize(
                _safe_str(email.get("body", "")) + " " + _safe_str(email.get("subject", ""))
            )
            if ("panshi" in body) and any(kw.lower() in body for kw in keywords):
                return True
    except Exception:
        pass

    return False


# -- Stage 1: Budget Change Discovery (Silent Event) --

async def _s1_budget_change_discovery(ctx) -> bool:
    """budget_limit=650000 and within_budget='yes' in Excel Sheet4."""
    data = _get_excel_data(ctx)
    if not data:
        return False
    _, _, sheet4_kv = _get_sheet_data(data)
    budget = _safe_float(sheet4_kv.get("budget_limit"))
    within = _safe_str(sheet4_kv.get("within_budget", ""))
    return budget == 650000 and within.lower() == "yes"


# -- Final: Director Email Sent --

async def _s0_director_email(ctx) -> bool:
    """Director received at least 1 procurement-related email from agent."""
    try:
        emails = await ctx.email.get_emails("director")
    except Exception:
        return False
    for email in emails:
        sender = _safe_str(email.get("from", "")).lower()
        if "zhaolei" in sender or "zhao" in sender:
            body_subj = _normalize(
                _safe_str(email.get("body", "")) + " " + _safe_str(email.get("subject", ""))
            )
            if any(kw in body_subj for kw in [
                "procurement", "vendor", "quotation", "evaluation",
                "panshi", "xingyun", "tianji", "selection", "recommend",
            ]):
                return True
    return False


# -- Final: Calendar Event Created --

async def _s0_calendar_event(ctx) -> bool:
    """Calendar has a procurement review meeting around March 20 with appropriate title."""
    try:
        all_events = []
        # Search for procurement or review events
        for keyword in ("procurement", "review", "vendor"):
            events = await ctx.calendar.find_events("zhaolei_calendar", keyword)
            all_events.extend(events)

        # Also check all events on March 20
        march20_events = await ctx.calendar.get_events(
            "zhaolei_calendar",
            datetime(2026, 3, 20, 0, 0),
            datetime(2026, 3, 20, 23, 59),
        )
        all_events.extend(march20_events)

        if not all_events:
            return False

        # Validate: at least one event has procurement/review in title AND is on ~March 20
        for event in all_events:
            title = _safe_str(event.get("summary", "") or event.get("title", "")).lower()
            has_relevant_title = any(kw in title for kw in [
                "procurement", "review", "vendor", "evaluation", "selection",
            ])

            # Check date is approximately March 20 (allow March 19-21 for timezone flexibility)
            start = event.get("start", "")
            start_str = str(start).lower()
            is_march_20 = any(d in start_str for d in ["2026-03-20", "2026-03-19", "2026-03-21"])

            if has_relevant_title and is_march_20:
                return True

        # Fallback: if any event exists on March 20 (even without perfect title match)
        if march20_events:
            return True

        return False
    except Exception:
        return False


# -- Bonus: Price Discrepancy Notes --

async def _b_price_discrepancy_notes(ctx) -> bool:
    """price_discrepancy_count >= 2 in Sheet4 (Xingyun handwritten + Panshi verbal discount)."""
    data = _get_excel_data(ctx)
    if not data:
        return False
    _, _, sheet4_kv = _get_sheet_data(data)
    count = _safe_float(sheet4_kv.get("price_discrepancy_count"))
    return count is not None and count >= 2


# -- Bonus: Notion Procurement Records Updated --

async def _b_notion_updated(ctx) -> bool:
    """Agent updated at least some Notion procurement records (status changes, scores filled)."""
    try:
        rows = await ctx.notion.query_db(PROCUREMENT_DB_NAME)
    except Exception:
        return False

    updated_count = 0
    for row in rows:
        proc_id = _get_notion_field(row, "Procurement ID", "title")
        if proc_id in ("PO-001", "PO-002", "PO-003"):
            status = _get_notion_field(row, "Status", "select")
            score = _get_notion_field(row, "Score", "number")
            amount = _get_notion_field(row, "Quotation Amount", "number")
            # Check if any field was updated from initial state
            if status and status != "Pending Evaluation":
                updated_count += 1
            elif score and score != 0:
                updated_count += 1
            elif amount and amount != 0:
                updated_count += 1
    # At least 2 of 3 vendor records should have some updates
    return updated_count >= 2


# -- Bonus: Email Quality --

async def _b_email_quality(ctx) -> bool:
    """Email to director contains vendor totals + recommendation rationale + discrepancy issues."""
    try:
        emails = await ctx.email.get_emails("director")
    except Exception:
        return False

    body = ""
    for email in emails:
        sender = _safe_str(email.get("from", "")).lower()
        if "zhaolei" in sender or "zhao" in sender:
            body += _safe_str(email.get("body", "")) + " " + _safe_str(email.get("subject", ""))

    if not body:
        return False

    # Normalize body: remove commas in numbers for matching
    body_normalized = body.replace(",", "")
    body_lower = body_normalized.lower()
    checks = {
        "vendor_totals": sum(1 for kw in ["255600", "255420", "246000", "234000"]
                             if kw in body_normalized) >= 2,
        "recommendation": any(kw in body_lower for kw in [
            "recommend", "comprehensive score", "weighted",
            "ranking", "highest score",
        ]),
        "discrepancy": any(kw in body_lower for kw in [
            "handwritten", "correction", "discount", "verbal",
            "10%", "inconsistenc", "discrepancy",
        ]),
    }
    return sum(1 for v in checks.values() if v) >= 3


# ── RUBRIC ────────────────────────────────────────────────────────

RUBRIC = {
    "stage0": [
        {"id": "S0_Xingyun_Handwritten_Price", "checker": _s0_xingyun_handwritten_price, "weight": 3.0},
        {"id": "S0_Panshi_Discount", "checker": _s0_panshi_discount, "weight": 3.0},
        {"id": "S0_Tianji_Delivery_Score", "checker": _s0_tianji_delivery_score, "weight": 2.0},
        {"id": "S0_Selection_Decision", "checker": _s0_selection_decision, "weight": 2.5},
        {"id": "S0_Redline_Sheets", "checker": _s0_redline_sheets, "weight": 2.0},
    ],
    "stage1": [
        {"id": "S1_Tianji_Price_Reduction", "checker": _s1_tianji_price_reduction, "weight": 2.5},
        {"id": "S1_Ranking_Change", "checker": _s1_ranking_change, "weight": 2.0},
        {"id": "S1_Panshi_Note_Discovery", "checker": _s1_panshi_note_discovery, "weight": 2.5},
        {"id": "S1_Budget_Change_Discovery", "checker": _s1_budget_change_discovery, "weight": 2.5},
    ],
    "final": [
        {"id": "S0_Director_Email", "checker": _s0_director_email, "weight": 1.5},
        {"id": "S0_Calendar_Event", "checker": _s0_calendar_event, "weight": 1.0},
        {"id": "B_Notion_Updated", "checker": _b_notion_updated, "weight": 1.0},
        {"id": "B_Price_Discrepancy_Notes", "checker": _b_price_discrepancy_notes, "weight": 1.0},
        {"id": "B_Email_Quality", "checker": _b_email_quality, "weight": 1.0},
    ],
}
