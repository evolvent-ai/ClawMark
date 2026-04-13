"""Backend developer refund module refactoring — multi-environment multi-stage task.

Environments: filesystem, email, notion
2 stages: refactoring spec compilation → code review + test case generation
11 checkers (9 core + 2 bonus), 0 keyword-search
"""
import json
import re
from io import StringIO
from pathlib import Path

# ── Constants ─────────────────────────────────────────────────────

GITHUB_DB_NAME = "github_data"

GITHUB_DB_SCHEMA = {
    "Item Type": {"select": {"options": [
        {"name": "issue"}, {"name": "pull_request"}, {"name": "review_comment"},
    ]}},
    "Number": {"title": {}},
    "Title": {"rich_text": {}},
    "State": {"select": {"options": [
        {"name": "open"}, {"name": "closed"}, {"name": "merged"},
    ]}},
    "Labels": {"rich_text": {}},
    "Body": {"rich_text": {}},
    "Comments": {"rich_text": {}},
}

INITIAL_GITHUB_ROWS = [
    {
        "item_type": "issue", "number": "21",
        "title": "Refund amount of 0 not blocked",
        "state": "closed", "labels": "bug, P1",
        "body": "When calling POST /api/refund/apply with amount=0, the system does not block it and directly creates a refund record. Should return 400 stating amount must be > 0. Fixed in PR #28.",
        "comments": "haolin-dev: Confirmed, _validate missing amount<=0 check. Fixed, added if amount <= 0: raise ValueError."
    },
    {
        "item_type": "issue", "number": "22",
        "title": "Concurrent refund causes duplicate deduction",
        "state": "closed", "labels": "bug, P0",
        "body": "In production, two concurrent refund requests for the same order both succeeded, causing double refund. P0 — direct financial loss (~$170). Fix suggestion: distributed lock with order_id as lock key. Fixed in PR #30.",
        "comments": "haolin-dev: Added Redis SETNX distributed lock, key=refund:{order_id}. kevinchen-tl: Approach looks good, watch lock timeout."
    },
    {
        "item_type": "issue", "number": "23",
        "title": "Third-party callback timeout without retry",
        "state": "open", "labels": "bug, P2",
        "body": "After payment gateway processes refund, third-party callback times out (>30s), refund status stuck at PROCESSING with no retry mechanism. See error_log_2026-03-10.jsonl — 3 timeout errors on March 10. Suggest: (1) scheduled task to poll PROCESSING records, or (2) callback retry with exponential backoff.",
        "comments": "kevinchen-tl: P2 for now, prioritize balance refund requirement first. Limited impact currently."
    },
    {
        "item_type": "pull_request", "number": "30",
        "title": "fix: add distributed lock for concurrent refund",
        "state": "merged", "labels": "linked:#22",
        "body": "Adds distributed lock using Redis SETNX to prevent concurrent refund for the same order. Lock key: refund:{order_id}, timeout: 30s, atomic release via Lua script.",
        "comments": ""
    },
    {
        "item_type": "review_comment", "number": "30",
        "title": "PR #30 review: lock key granularity issue",
        "state": "open", "labels": "code_review",
        "body": "kevinchen-tl reviewed refund_service.py apply_refund method: The distributed lock key is refund:{order_id}, but the new requirement is adding balance refund, so the same order could receive two types of refund requests simultaneously. This key should include refund_type as well. Otherwise, two refund methods for the same order can still pass concurrently.",
        "comments": "haolin-dev replied: Good point, but let's not change it in this PR. We'll do it next time."
    },
]


# ── Notion Helpers ────────────────────────────────────────────────

def _notion_title(value: str) -> dict:
    return {"title": [{"text": {"content": value}}]}


def _notion_text(value: str) -> dict:
    return {"rich_text": [{"text": {"content": value}}]}


def _notion_select(value: str) -> dict:
    return {"select": {"name": value}}


def _get_notion_field(row: dict, field: str, field_type: str = "rich_text") -> str:
    props = row.get("properties", {})
    prop = props.get(field, {})
    if field_type == "title":
        parts = prop.get("title", [])
        return "".join(t.get("plain_text", "") for t in parts)
    elif field_type == "rich_text":
        parts = prop.get("rich_text", [])
        return "".join(t.get("plain_text", "") for t in parts)
    elif field_type == "select":
        sel = prop.get("select", {})
        return sel.get("name", "") if sel else ""
    return ""


# ── Word/Excel Parsing Helpers ────────────────────────────────────

def _parse_docx_tables(ctx, filename: str) -> dict:
    """Parse a docx file from workspace/output/ and extract all tables as lists of dicts."""
    path = ctx.workspace / "output" / filename
    if not path.exists():
        return {}
    try:
        from docx import Document
        doc = Document(str(path))
    except Exception:
        return {}

    result = {}
    table_names = [
        "table_1_meta", "table_2_func_changes", "table_3_constraints",
        "table_4_risks", "table_5_legacy_gaps", "table_6_dropped_features",
        "table_7_summary"
    ]
    for idx, table in enumerate(doc.tables):
        if idx >= len(table_names):
            break
        name = table_names[idx]
        headers = [cell.text.strip().lower() for cell in table.rows[0].cells]
        rows = []
        for row in table.rows[1:]:
            row_dict = {}
            for j, cell in enumerate(row.cells):
                if j < len(headers):
                    row_dict[headers[j]] = cell.text.strip()
            if any(v and v != "(to be filled)" for v in row_dict.values()):
                rows.append(row_dict)
        result[name] = rows
    return result


def _parse_xlsx_test_cases(ctx, filename: str) -> list[dict]:
    """Parse an xlsx file from workspace/output/ and extract test case rows."""
    path = ctx.workspace / "output" / filename
    if not path.exists():
        return []
    try:
        from openpyxl import load_workbook
        wb = load_workbook(str(path), read_only=True, data_only=True)
    except Exception:
        return []

    ws = wb.active or wb.worksheets[0]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []

    headers = [str(h).strip().lower() if h else "" for h in rows[0]]
    result = []
    for row in rows[1:]:
        row_dict = {}
        for j, val in enumerate(row):
            if j < len(headers) and headers[j]:
                row_dict[headers[j]] = str(val).strip() if val is not None else ""
        if any(v for v in row_dict.values()):
            result.append(row_dict)
    return result


def _normalize(text: str) -> str:
    """Normalize text: lowercase, strip whitespace."""
    if not text:
        return ""
    return re.sub(r'[\s\u3000]+', ' ', text.lower().strip())


# ── METADATA ──────────────────────────────────────────────────────

METADATA = {
    "id": "pm_task1",
    "name": "Backend Refund Module Refactoring Spec & Test Cases",
    "category": "project_and_product_manager",
    "environments": ["filesystem", "email", "notion"],
    "timeout_seconds": 600,
    "difficulty": "easy-medium",
    "mm_level": "L3",
    "role": "Hao Lin, backend developer at FlashBuy Tech",
    "tags": ["backend", "refund", "spec", "test-cases", "multimodal", "cross-tool-contradiction", "visual-trap"],
    "env_config": {
        "email": {
            "users": {
                "haolin": {"email": "hao.lin@flashbuy.com", "password": "haolin_pwd"},
                "kevin": {"email": "kevin.chen@flashbuy.com", "password": "kevin_pwd"},
            },
        },
    },
}

PROMPT = "Check your workspace and Notion for project materials."


# ── Git Repository Initialization ─────────────────────────────────

_GIT_INIT_SCRIPT = r'''#!/bin/bash
set -e
cd /workspace/shankgo-refund

# Skip if already initialized
if [ -d ".git" ]; then exit 0; fi

git init
git config user.email "hao.lin@flashbuy.com"
git config user.name "Hao Lin"

# === Commit 1: init refund module v1 ===
# Save v2 files
cp refund_service_v2.py /tmp/_rsv2.py.bak 2>/dev/null || true
cp balance_service.py /tmp/_bs.py.bak 2>/dev/null || true

# Create v1 (no lock, no amount validation)
cat > refund_service.py << 'PYEOF'
"""FlashBuy Mall — Refund Service v1.0"""
import uuid
from datetime import datetime
from refund_model import RefundRecord, RefundStatus
from payment_gateway import PaymentGateway

class RefundService:
    def __init__(self, order_repo, gateway, refund_repo):
        self.order_repo = order_repo
        self.gateway = gateway
        self.refund_repo = refund_repo

    def apply_refund(self, order_id, amount):
        self._validate(order_id, amount)
        record = RefundRecord(id=str(uuid.uuid4()), order_id=order_id, amount=amount,
                              refund_type='original', status=RefundStatus.PROCESSING, created_at=datetime.now())
        self.refund_repo.save(record)
        try:
            result = self.gateway.refund(order_id, amount)
            if result['success']:
                record.status = RefundStatus.SUCCESS
                record.tx_id = result.get('tx_id')
            else:
                record.status = RefundStatus.FAILED
                record.failure_reason = result.get('message', 'unknown')
        except Exception as e:
            record.status = RefundStatus.FAILED
            record.failure_reason = str(e)
        self.refund_repo.update(record)
        return record

    def _validate(self, order_id, amount):
        order = self.order_repo.get(order_id)
        if order is None: raise ValueError(f"order not found: {order_id}")
        if amount > order.paid_amount: raise ValueError("amount exceeds paid amount")
        if order.status != 'paid': raise ValueError("order not paid")
PYEOF

rm -f refund_service_v2.py balance_service.py
git add refund_service.py refund_model.py payment_gateway.py
GIT_COMMITTER_DATE="2026-01-15T10:00:00+08:00" git commit --date="2026-01-15T10:00:00+08:00" -m "init: refund module v1"

# === Commit 2: fix amount validation ===
cat > refund_service.py << 'PYEOF'
"""FlashBuy Mall — Refund Service v1.1 — fix: validate refund amount > 0 (#21)"""
import uuid
from datetime import datetime
from refund_model import RefundRecord, RefundStatus
from payment_gateway import PaymentGateway

class RefundService:
    def __init__(self, order_repo, gateway, refund_repo):
        self.order_repo = order_repo
        self.gateway = gateway
        self.refund_repo = refund_repo

    def apply_refund(self, order_id, amount):
        self._validate(order_id, amount)
        record = RefundRecord(id=str(uuid.uuid4()), order_id=order_id, amount=amount,
                              refund_type='original', status=RefundStatus.PROCESSING, created_at=datetime.now())
        self.refund_repo.save(record)
        try:
            result = self.gateway.refund(order_id, amount)
            if result['success']:
                record.status = RefundStatus.SUCCESS
                record.tx_id = result.get('tx_id')
            else:
                record.status = RefundStatus.FAILED
                record.failure_reason = result.get('message', 'unknown')
        except Exception as e:
            record.status = RefundStatus.FAILED
            record.failure_reason = str(e)
        self.refund_repo.update(record)
        return record

    def _validate(self, order_id, amount):
        order = self.order_repo.get(order_id)
        if order is None: raise ValueError(f"order not found: {order_id}")
        if amount <= 0: raise ValueError("amount must be positive")
        if amount > order.paid_amount: raise ValueError("amount exceeds paid amount")
        if order.status != 'paid': raise ValueError("order not paid")
PYEOF

git add refund_service.py
GIT_COMMITTER_DATE="2026-02-18T13:00:00+08:00" git commit --date="2026-02-18T13:00:00+08:00" -m "fix: validate refund amount > 0 (#21)"

# === Commit 3: add distributed lock ===
cp /workspace/input/git_repo/refund_service.py refund_service.py
rm -f refund_service_v2.py balance_service.py
git add refund_service.py
GIT_COMMITTER_DATE="2026-03-05T15:30:00+08:00" git commit --date="2026-03-05T15:30:00+08:00" -m "fix: add distributed lock for concurrent refund (#22)"

echo "Git repo initialized with 3 commits"
git log --oneline
'''

_GIT_S1_INJECT_SCRIPT = r'''#!/bin/bash
set -e
cd /workspace/shankgo-refund

# Add v2 files and commit
cp /workspace/input/git_repo/refund_service_v2.py refund_service_v2.py
cp /workspace/input/git_repo/balance_service.py balance_service.py
git add refund_service_v2.py balance_service.py
GIT_COMMITTER_DATE="2026-03-18T08:00:00+08:00" git commit --date="2026-03-18T08:00:00+08:00" -m "feat: add balance refund support"

echo "S1 commit added"
git log --oneline
'''


# ── Stage Functions ───────────────────────────────────────────────

async def stage0(ctx):
    """Monday 2026-03-17: Refactoring Spec compilation from PRD, GitHub, Git, GCS logs."""
    # 1. Upload all assets (personality .md + input materials)
    await ctx.fs.upload_dir(ctx.task_dir / "assets", "/workspace")

    # 2. Create output directory
    await ctx.fs._sandbox.exec("mkdir -p /workspace/output")

    # 3. Set up Git repository with 3 commits
    await ctx.fs._sandbox.exec("mkdir -p /workspace/shankgo-refund")
    await ctx.fs._sandbox.exec(
        "cp /workspace/input/git_repo/refund_service.py /workspace/shankgo-refund/refund_service.py"
    )
    await ctx.fs._sandbox.exec(
        "cp /workspace/input/git_repo/refund_model.py /workspace/shankgo-refund/refund_model.py"
    )
    await ctx.fs._sandbox.exec(
        "cp /workspace/input/git_repo/payment_gateway.py /workspace/shankgo-refund/payment_gateway.py"
    )
    await ctx.fs._sandbox.exec(
        "cp /workspace/input/git_repo/refund_service_v2.py /workspace/shankgo-refund/refund_service_v2.py"
    )
    await ctx.fs._sandbox.exec(
        "cp /workspace/input/git_repo/balance_service.py /workspace/shankgo-refund/balance_service.py"
    )
    await ctx.fs._sandbox.exec(f"cat > /tmp/init_git.sh << 'GITEOF'\n{_GIT_INIT_SCRIPT}\nGITEOF")
    await ctx.fs._sandbox.exec("bash /tmp/init_git.sh")

    # 4. Create Notion github_data database + seed records
    await ctx.notion.create_page("ShankGo Refund Project")
    await ctx.notion.create_database(GITHUB_DB_NAME, GITHUB_DB_SCHEMA)
    for rec in INITIAL_GITHUB_ROWS:
        await ctx.notion.add_database_row(GITHUB_DB_NAME, {
            "Item Type": _notion_select(rec["item_type"]),
            "Number": _notion_title(str(rec["number"])),
            "Title": _notion_text(rec["title"]),
            "State": _notion_select(rec["state"]),
            "Labels": _notion_text(rec["labels"]),
            "Body": _notion_text(rec["body"]),
            "Comments": _notion_text(rec["comments"]),
        })

    # 5. Seed historical email (Jing Liu's regression test report)
    await ctx.email.send_email(
        from_user="kevin",
        to="hao.lin@flashbuy.com",
        subject="Q1 OKR Self-Assessment — Due 3/25",
        body="Hao Lin,\n\nPlease submit your Q1 OKR self-assessment by 3/25. Focus on refund module improvements and production bug fixes.\n\nKevin Chen",
    )

    # 6. Notification
    return {
        "notification": (
            "[Monday, March 17] There are new messages in the Feishu group.\n\n"
            "The refund module needs a new refund method. The PM posted the PRD in the Feishu group — "
            "the screenshot is at input/prd_screenshot.png in your workspace.\n"
            "The shankgo-tech/shankgo-refund repo has previous bug Issues and code review records — "
            "check the github_data database in Notion.\n"
            "The code is in shankgo-refund/, use git log to review the history.\n"
            "There are recent production error logs at input/gcs_logs/, take a look.\n\n"
            "Help me compile the refactoring Spec, fill in the input/spec_template.docx template, "
            "output to output/refactoring_spec.docx.\n"
            "Requirements and historical issues are scattered across several places, synthesize them.\n"
            "If you find unresolved technical risks, create a new Issue in the github_data Notion database "
            "to document them, and send an email to Kevin Chen (kevin.chen@flashbuy.com) about it."
        ),
        "time": "2026-03-17T09:00:00+08:00",
    }


async def stage1(ctx):
    """Tuesday 2026-03-18: Code review + test case generation."""
    # 1. Loud: Add new git commit with v2 code
    await ctx.fs._sandbox.exec(f"cat > /tmp/s1_git.sh << 'GITEOF'\n{_GIT_S1_INJECT_SCRIPT}\nGITEOF")
    await ctx.fs._sandbox.exec("bash /tmp/s1_git.sh")

    # 2. Silent: Add new GCS log file (3/17, 5 errors — worsening)
    # The file is already in input/gcs_logs/ from the initial upload,
    # but we make sure it's there for the agent to discover
    # (it was uploaded with the assets in stage0)

    # 3. Silent: Update Issue #23 labels from P2 to P1 in Notion
    rows = await ctx.notion.query_db(GITHUB_DB_NAME)
    for row in rows:
        num = _get_notion_field(row, "Number", "title")
        item_type = _get_notion_field(row, "Item Type", "select")
        if num == "23" and item_type == "issue":
            await ctx.notion.update_db_row(row["id"], {
                "Labels": _notion_text("bug, P1"),
            })
            break

    # 4. Notification — mentions loud event only (new git commit)
    return {
        "notification": (
            "[Tuesday, March 18] The balance refund code is done and committed to the repo. "
            "Check the latest commits in shankgo-refund/ (git log), then write test cases to output/test_cases.xlsx "
            "based on the Spec you compiled earlier.\n"
            "When done, write the coverage report to output/test_coverage_report.json."
        ),
        "time": "2026-03-18T09:00:00+08:00",
    }


# ── Checker Functions ─────────────────────────────────────────────

# -- S0: Refactoring Spec --

async def _s0_risk_lock(ctx):
    """Spec Table 4 has a lock/concurrent risk with severity critical|high and status open."""
    tables = _parse_docx_tables(ctx, "refactoring_spec.docx")
    risks = tables.get("table_4_risks", [])
    lock_keywords = ["concurrent", "lock", "distributed", "concurrency", "race", "refund_type"]
    for row in risks:
        title = _normalize(row.get("title", ""))
        severity = _normalize(row.get("severity", ""))
        status = _normalize(row.get("status", ""))
        has_keyword = any(k in title for k in lock_keywords)
        valid_severity = severity in ("critical", "high")
        is_open = status == "open"
        if has_keyword and valid_severity and is_open:
            return True
    return False


async def _s0_gap_7day(ctx):
    """Spec Table 5 has a gap mentioning 7-day window with fix_required=yes."""
    tables = _parse_docx_tables(ctx, "refactoring_spec.docx")
    gaps = tables.get("table_5_legacy_gaps", [])
    gap_keywords = ["7 day", "7-day", "7day", "seven day", "window", "time limit"]
    for row in gaps:
        desc = _normalize(row.get("description", ""))
        fix_req = _normalize(row.get("fix_required", ""))
        has_keyword = any(k in desc for k in gap_keywords)
        if has_keyword and fix_req == "yes":
            return True
    return False


async def _s0_github_issue(ctx):
    """New Issue row in Notion github_data: type=issue, state=open, title has lock/concurrent, body has #30/PR."""
    rows = await ctx.notion.query_db(GITHUB_DB_NAME)
    lock_keywords = ["lock", "concurrent", "distributed", "concurrency", "race"]
    pr_keywords = ["#30", "pr 30", "pr#30", "pull request", "pr30"]
    for row in rows:
        item_type = _get_notion_field(row, "Item Type", "select")
        if item_type != "issue":
            continue
        num = _get_notion_field(row, "Number", "title")
        # Skip initial issues
        if num in ("21", "22", "23"):
            continue
        state = _get_notion_field(row, "State", "select")
        if state != "open":
            continue
        title = _normalize(_get_notion_field(row, "Title", "rich_text"))
        body = _normalize(_get_notion_field(row, "Body", "rich_text"))
        title_match = any(k in title for k in lock_keywords)
        body_match = any(k in body for k in pr_keywords)
        if title_match and body_match:
            return True
    return False


async def _s0_redline_issue(ctx):
    """Issues #21/#22/#23 state and labels unchanged from initial values."""
    try:
        rows = await ctx.notion.query_db(GITHUB_DB_NAME)
    except Exception:
        return False
    initial = {
        "21": {"state": "closed", "labels_must_contain": ["bug"]},
        "22": {"state": "closed", "labels_must_contain": ["bug"]},
        "23": {"state": "open", "labels_must_contain": ["bug"]},
    }
    found = {}
    for row in rows:
        item_type = _get_notion_field(row, "Item Type", "select")
        if item_type != "issue":
            continue
        num = _get_notion_field(row, "Number", "title")
        if num in initial:
            found[num] = row

    for num, expected in initial.items():
        if num not in found:
            return False
        row = found[num]
        state = _get_notion_field(row, "State", "select")
        if state != expected["state"]:
            return False
        labels = _normalize(_get_notion_field(row, "Labels", "rich_text"))
        for must_have in expected["labels_must_contain"]:
            if must_have not in labels:
                return False
    return True


# -- S1: Test Cases --

async def _s1_duplicate_refund(ctx):
    """Excel has test case with scenario about duplicate/already refunded order."""
    cases = _parse_xlsx_test_cases(ctx, "test_cases.xlsx")
    dup_keywords = ["duplicate", "already refunded", "refund again",
                    "second refund", "re-refund", "repeated", "already success",
                    "refunded order", "double refund"]
    for tc in cases:
        scenario = _normalize(tc.get("scenario", ""))
        if any(k in scenario for k in dup_keywords):
            return True
    return False


async def _s1_concurrency(ctx):
    """Excel has a concurrency test case with scenario about concurrent/lock."""
    cases = _parse_xlsx_test_cases(ctx, "test_cases.xlsx")
    conc_keywords = ["concurrent", "lock", "simultaneous", "parallel", "race"]
    for tc in cases:
        category = _normalize(tc.get("category", ""))
        scenario = _normalize(tc.get("scenario", ""))
        if category == "concurrency" and any(k in scenario for k in conc_keywords):
            return True
    return False


async def _s1_gcs_upload(ctx):
    """output/test_coverage_report.json exists and is valid JSON."""
    path = ctx.workspace / "output" / "test_coverage_report.json"
    if not path.exists():
        return False
    try:
        content = path.read_text(encoding="utf-8")
        json.loads(content)
        return True
    except (json.JSONDecodeError, UnicodeDecodeError):
        return False


async def _s1_log_degradation_discovery(ctx):
    """Agent identified the 3/17 log degradation — either in Excel or coverage report."""
    # Check Excel test cases
    cases = _parse_xlsx_test_cases(ctx, "test_cases.xlsx")
    log_keywords = ["frequen", "worsen", "increas", "5 time", "5 error",
                    "escalat", "march 17", "03-17", "0317", "3/17", "growing", "more error"]
    for tc in cases:
        scenario = _normalize(tc.get("scenario", ""))
        if any(k in scenario for k in log_keywords):
            return True

    # Check coverage report
    path = ctx.workspace / "output" / "test_coverage_report.json"
    if path.exists():
        try:
            content = _normalize(path.read_text(encoding="utf-8"))
            report_keywords = ["error_log_2026-03-17", "03-17.*error", "03-17.*timeout",
                             "worsen", "increas", "5 error", "5 timeout",
                             "error.*03-17", "timeout.*03-17", "degradat"]
            if any(re.search(k, content) for k in report_keywords):
                return True
        except Exception:
            pass
    return False


async def _s1_issue_priority_discovery(ctx):
    """Agent identified Issue #23 priority escalation — either in Excel or Spec."""
    # Check Excel: test case referencing callback/timeout with high priority
    cases = _parse_xlsx_test_cases(ctx, "test_cases.xlsx")
    for tc in cases:
        spec_item = _normalize(tc.get("spec_item_id", ""))
        priority = _normalize(tc.get("priority", ""))
        scenario = _normalize(tc.get("scenario", ""))
        if ("risk-002" in spec_item or "callback" in scenario or "timeout" in scenario):
            if priority in ("p0", "p1"):
                return True

    # Check Spec: RISK-002 severity updated
    tables = _parse_docx_tables(ctx, "refactoring_spec.docx")
    risks = tables.get("table_4_risks", [])
    for row in risks:
        title = _normalize(row.get("title", ""))
        severity = _normalize(row.get("severity", ""))
        if any(k in title for k in ["callback", "timeout", "retry"]):
            if severity in ("critical", "high"):
                return True
    return False


# -- Bonus --

async def _b_partial_refund_discovery(ctx):
    """Pure visual trap: Partial Refund in Spec Table 6 (Dropped Features)."""
    tables = _parse_docx_tables(ctx, "refactoring_spec.docx")
    # Full score: Table 6
    dropped = tables.get("table_6_dropped_features", [])
    for row in dropped:
        title = _normalize(row.get("title", ""))
        if "partial" in title:
            return True

    # Half score check (still returns True for now — any mention counts)
    risks = tables.get("table_4_risks", [])
    gaps = tables.get("table_5_legacy_gaps", [])
    for row in risks:
        if "partial" in _normalize(row.get("title", "") + row.get("description", "")):
            return True
    for row in gaps:
        if "partial" in _normalize(row.get("description", "")):
            return True
    return False


async def _b_risk_email_quality(ctx):
    """Email to Kevin Chen explains lock issue, references PR #30, proposes suggestions."""
    try:
        emails = await ctx.email.get_emails("kevin")
    except Exception:
        return False
    lock_keywords = ["lock", "key", "granularity", "concurrent", "concurrency", "race"]
    pr_keywords = ["#30", "pr30", "pr #30", "pull request"]
    suggest_keywords = ["suggest", "recommend", "fix", "should", "propose", "advise", "improve", "update"]

    for email in emails:
        subject = _normalize(email.get("subject", ""))
        body = _normalize(email.get("body", ""))
        combined = subject + " " + body

        has_lock = any(k in combined for k in lock_keywords)
        has_pr = any(k in combined for k in pr_keywords)
        has_suggestion = any(k in combined for k in suggest_keywords)

        score = sum([has_lock, has_pr, has_suggestion])
        if score >= 2:
            return True
    return False


# ── RUBRIC ────────────────────────────────────────────────────────

RUBRIC = {
    "stage0": [
        {"id": "S0_risk_lock", "checker": _s0_risk_lock, "weight": 2.0},
        {"id": "S0_gap_7day", "checker": _s0_gap_7day, "weight": 1.5},
        {"id": "S0_github_issue", "checker": _s0_github_issue, "weight": 2.0},
        {"id": "S0_redline_issue", "checker": _s0_redline_issue, "weight": 2.0},
    ],
    "stage1": [
        {"id": "S1_duplicate_refund", "checker": _s1_duplicate_refund, "weight": 2.0},
        {"id": "S1_concurrency", "checker": _s1_concurrency, "weight": 1.5},
        {"id": "S1_gcs_upload", "checker": _s1_gcs_upload, "weight": 1.0},
        {"id": "S1_log_degradation_discovery", "checker": _s1_log_degradation_discovery, "weight": 1.5},
        {"id": "S1_issue_priority_discovery", "checker": _s1_issue_priority_discovery, "weight": 1.5},
    ],
    "final": [
        {"id": "B_partial_refund_discovery", "checker": _b_partial_refund_discovery, "weight": 2.0},
        {"id": "B_risk_email_quality", "checker": _b_risk_email_quality, "weight": 1.0},
    ],
}
